from __future__ import annotations

import argparse
import csv
import json
import sqlite3
from pathlib import Path
from typing import Any, Iterable

ROOT = Path(__file__).resolve().parents[1]
REQUIRED = {
    "settlement_main": ["setl_id", "mdtrt_id", "fixmedins_code", "psn_no", "gend", "age", "fund_pay_sumamt", "setl_time"],
    "fee_detail": ["setl_id", "mdtrt_id", "hilist_code", "hilist_name", "cnt", "pric", "det_item_fee_sumamt", "fee_ocur_time"],
}
NUMERIC_FIELDS = {"age", "fund_pay_sumamt", "cnt", "pric", "det_item_fee_sumamt"}


def resolve_path(path_text: str) -> Path:
    path = Path(path_text)
    if path.is_absolute():
        return path
    return ROOT / path


def clean_header(value: Any) -> str:
    return str(value or "").replace("\ufeff", "").strip()


def normalize_value(field: str, value: Any) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    if field in NUMERIC_FIELDS:
        text = text.replace(",", "")
        try:
            return float(text)
        except ValueError:
            return None
    return text


def load_csv_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = [clean_header(item) for item in (reader.fieldnames or [])]
        rows = []
        for row in reader:
            rows.append({clean_header(key): value for key, value in row.items()})
    return headers, rows


def load_xlsx_rows(path: Path, sheet: str | None) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to import xlsx files") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [clean_header(item) for item in next(rows_iter)]
    rows = []
    for values in rows_iter:
        rows.append({headers[index]: values[index] if index < len(values) else None for index in range(len(headers))})
    return headers, rows


def load_rows(path: Path, sheet: str | None) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx_rows(path, sheet)
    raise ValueError(f"unsupported file type: {path}")


def map_columns(table: str, headers: list[str], table_config: dict[str, Any]) -> dict[str, str]:
    header_lookup = {clean_header(header): clean_header(header) for header in headers}
    mapping: dict[str, str] = {}
    configured = table_config.get("columns", {})
    for target in REQUIRED[table]:
        aliases = configured.get(target, [target])
        for alias in aliases:
            clean_alias = clean_header(alias)
            if clean_alias in header_lookup:
                mapping[target] = header_lookup[clean_alias]
                break
        if target not in mapping:
            raise ValueError(f"{table} missing required column for {target}; known headers: {headers}")
    return mapping


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        DROP TABLE IF EXISTS settlement_main;
        DROP TABLE IF EXISTS fee_detail;

        CREATE TABLE settlement_main (
          setl_id TEXT PRIMARY KEY,
          mdtrt_id TEXT,
          fixmedins_code TEXT,
          psn_no TEXT,
          gend TEXT,
          age REAL,
          fund_pay_sumamt REAL,
          setl_time TEXT
        );

        CREATE TABLE fee_detail (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          setl_id TEXT,
          mdtrt_id TEXT,
          hilist_code TEXT,
          hilist_name TEXT,
          cnt REAL,
          pric REAL,
          det_item_fee_sumamt REAL,
          fee_ocur_time TEXT
        );

        CREATE INDEX idx_settlement_main_mdtrt_id ON settlement_main(mdtrt_id);
        CREATE INDEX idx_settlement_main_fixmedins_code ON settlement_main(fixmedins_code);
        CREATE INDEX idx_fee_detail_setl_id ON fee_detail(setl_id);
        CREATE INDEX idx_fee_detail_hilist_code ON fee_detail(hilist_code);
        CREATE INDEX idx_fee_detail_fee_ocur_time ON fee_detail(fee_ocur_time);
        """
    )


def insert_rows(conn: sqlite3.Connection, table: str, rows: Iterable[dict[str, Any]], mapping: dict[str, str]) -> int:
    fields = REQUIRED[table]
    placeholders = ", ".join(["?"] * len(fields))
    sql = f"INSERT OR REPLACE INTO {table} ({', '.join(fields)}) VALUES ({placeholders})"
    count = 0
    batch = []
    for row in rows:
        batch.append(tuple(normalize_value(field, row.get(mapping[field])) for field in fields))
        if len(batch) >= 1000:
            conn.executemany(sql, batch)
            count += len(batch)
            batch = []
    if batch:
        conn.executemany(sql, batch)
        count += len(batch)
    return count


def import_table(conn: sqlite3.Connection, table: str, config: dict[str, Any]) -> dict[str, Any]:
    path = resolve_path(config["file"])
    if not path.exists():
        raise FileNotFoundError(f"{table} source file not found: {path}")
    headers, rows = load_rows(path, config.get("sheet"))
    mapping = map_columns(table, headers, config)
    count = insert_rows(conn, table, rows, mapping)
    return {"table": table, "source": str(path), "rows": count, "column_mapping": mapping}


def write_local_database_config(db_path: Path, max_rows: int) -> None:
    config_path = ROOT / "config" / "database.local.json"
    config = {
        "enabled": True,
        "provider": "sqlite",
        "readonly": True,
        "sqlite_path": str(db_path).replace("\\", "/"),
        "timeout_seconds": 10,
        "max_rows": max_rows,
    }
    config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import settlement and fee detail exports into a readonly SQLite database.")
    parser.add_argument("--mapping", default="config/claims_import_mapping.example.json")
    parser.add_argument("--output", default="data/claims_import/output/claims_real.db")
    parser.add_argument("--write-local-config", action="store_true")
    parser.add_argument("--max-rows", type=int, default=500)
    args = parser.parse_args()

    mapping_path = resolve_path(args.mapping)
    output_path = resolve_path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    config = json.loads(mapping_path.read_text(encoding="utf-8-sig"))

    if output_path.exists():
        output_path.unlink()
    conn = sqlite3.connect(output_path)
    try:
        create_schema(conn)
        report = [
            import_table(conn, "settlement_main", config["settlement_main"]),
            import_table(conn, "fee_detail", config["fee_detail"]),
        ]
        conn.commit()
    finally:
        conn.close()

    if args.write_local_config:
        write_local_database_config(output_path, args.max_rows)

    print(json.dumps({"database": str(output_path), "tables": report}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
